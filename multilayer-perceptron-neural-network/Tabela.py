from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from string import ascii_uppercase
from pathlib import Path
from datetime import datetime
import os
from datetime import date

thin_border = Border(left=Side(style='thin'),right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
redFill = PatternFill(start_color="FFFF0000",
end_color="FFFF0000",
fill_type="solid")

arquivo_excel = Workbook()

def __CriarTabela__(nameArchive):
    planilha = arquivo_excel.active
    planilha.title = "MPC"
    planilha['A1'] = 'TOPOLOGIA INTERMEDIÁRIA'
    planilha['B1'] = 'ACURÁCIA TREINAMENTO'
    planilha['C1'] = 'EQM'
    planilha['D1'] = 'NUMERO EPOCAS'
    planilha['E1'] = 'TEMPO DE EXECUÇÃO'
    planilha['F1'] = 'ACURACIA TESTE'

    for col in planilha.columns:
         max_length = 0
         letter = get_column_letter(col[0].column)
         cell = col[0] # mesma coisa de fazewr cell in col
         ft = Font(color=colors.BLACK, bold=True)
         cell.font = ft
         cell.border = thin_border
         cell.fill = redFill
         try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
         except:
            pass
    arquivo_excel.save(nameArchive)
    return arquivo_excel

def inserirValores(planilha, name, A,B,C,D,E,F):
    planilha.active.append((A,B,C,D,E,F))
    arquivo_excel.save(name)